from __future__ import annotations

import os
import uuid
from pathlib import Path
from typing import List, Optional

import io

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.database import get_db, get_duckdb
from app.middleware.auth import get_current_user
from app.models.dataset import Dataset
from app.schemas.auth import AuthenticatedUser
from app.schemas.dataset import DatasetListItem, DatasetOut, DatasetSchema, PreviewResponse
from app.services.schema_discovery import discover_schema

router = APIRouter(prefix="/datasets", tags=["datasets"])

_ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv"}


# ── POST /datasets/upload ─────────────────────────────────────────────────────

@router.post("/upload", response_model=DatasetOut, status_code=status.HTTP_201_CREATED)
async def upload_dataset(
    file: UploadFile,
    db: AsyncSession = Depends(get_db),
    current_user: AuthenticatedUser = Depends(get_current_user),
) -> DatasetOut:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in _ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported file type '{suffix}'. Allowed: {_ALLOWED_EXTENSIONS}",
        )

    # 1. Persist file to disk
    dataset_id = uuid.uuid4()
    upload_dir = Path(settings.upload_dir)
    upload_dir.mkdir(parents=True, exist_ok=True)
    file_path = upload_dir / f"{dataset_id}{suffix}"

    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)

    try:
        # 2. Discover schema
        schema: DatasetSchema = discover_schema(str(file_path))

        # 3. Store data into DuckDB
        _load_into_duckdb(str(file_path), suffix, dataset_id, schema)

        # 4. Persist metadata to PostgreSQL
        dataset_name = Path(file.filename or "").stem
        source_type = suffix.lstrip(".")
        dataset = Dataset(
            id=dataset_id,
            name=dataset_name,
            source_type=source_type,
            schema_info=schema.model_dump(),
            row_count=schema.row_count,
            file_path=str(file_path),
            created_by=current_user.user_id,
        )
        db.add(dataset)
        await db.commit()
        await db.refresh(dataset)

        # 自动权限配置：检测伙伴运营字段，自动授权 + 配 RLS
        await _auto_configure_permissions(db, dataset_id, schema)

        return dataset  # type: ignore[return-value]

    except Exception as exc:
        # Clean up uploaded file on failure
        try:
            os.remove(file_path)
        except OSError:
            pass
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Failed to process file: {exc}",
        )


def _load_into_duckdb(
    file_path: str, suffix: str, dataset_id: uuid.UUID, schema: DatasetSchema
) -> None:
    """Load file data into a DuckDB table named dataset_{uuid_hex}."""
    table_name = f"dataset_{dataset_id.hex}"
    conn = get_duckdb()

    # Drop table if it already exists (re-upload scenario)
    conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')

    if suffix in (".xlsx", ".xls"):
        # Read with openpyxl then INSERT via DuckDB
        import openpyxl
        # Patch InlineFont to tolerate unknown XML attrs (e.g. 'name') in some Excel files
        from openpyxl.cell.text import InlineFont as _IF
        _orig_if_init = _IF.__init__
        def _tolerant_if_init(self, rFont=None, **kw):
            kw.pop('name', None)
            _orig_if_init(self, rFont=rFont, **kw)
        _IF.__init__ = _tolerant_if_init

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return

        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        data_rows = rows[1:]

        # Build CREATE TABLE from discovered schema
        col_defs = ", ".join(
            f'"{col.name}" VARCHAR' for col in schema.columns
        )
        conn.execute(f'CREATE TABLE "{table_name}" ({col_defs})')

        # Batch insert
        placeholders = ", ".join("?" * len(headers))
        insert_sql = f'INSERT INTO "{table_name}" VALUES ({placeholders})'
        for row in data_rows:
            padded = list(row) + [None] * max(0, len(headers) - len(row))
            str_row = [
                str(v) if v is not None and str(v).strip() != "" else None
                for v in padded[: len(headers)]
            ]
            conn.execute(insert_sql, str_row)

    elif suffix == ".csv":
        # DuckDB can read CSVs natively
        conn.execute(
            f'CREATE TABLE "{table_name}" AS SELECT * FROM read_csv_auto(?)',
            [file_path],
        )

    conn.close()


# ── GET /datasets ─────────────────────────────────────────────────────────────

@router.get("", response_model=List[DatasetListItem])
async def list_datasets(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: AuthenticatedUser = Depends(get_current_user),
) -> List[DatasetListItem]:
    offset = (page - 1) * page_size
    result = await db.execute(
        select(Dataset)
        .where(Dataset.is_active == True)  # noqa: E712
        .order_by(Dataset.created_at.desc())
        .offset(offset)
        .limit(page_size)
    )
    datasets = result.scalars().all()
    return [
        DatasetListItem(
            id=ds.id,
            name=ds.name,
            source_type=ds.source_type,
            row_count=ds.row_count,
            column_count=len(ds.schema_info.get("columns", [])),
            created_at=ds.created_at,
        )
        for ds in datasets
    ]


# ── GET /datasets/trash ──────────────────────────────────────────────────────

@router.get("/trash", response_model=List[DatasetListItem])
async def list_trash(
    db: AsyncSession = Depends(get_db),
    current_user: AuthenticatedUser = Depends(get_current_user),
) -> List[DatasetListItem]:
    """列出已软删除的数据集（仅 admin）。"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可查看回收站")
    result = await db.execute(
        select(Dataset).where(Dataset.is_active == False).order_by(Dataset.created_at.desc())  # noqa: E712
    )
    datasets = result.scalars().all()
    return [
        DatasetListItem(
            id=ds.id,
            name=ds.name,
            source_type=ds.source_type,
            row_count=ds.row_count,
            column_count=len(ds.schema_info.get("columns", [])),
            created_at=ds.created_at,
        )
        for ds in datasets
    ]


# ── GET /datasets/{id} ────────────────────────────────────────────────────────

@router.get("/{dataset_id}", response_model=DatasetOut)
async def get_dataset(
    dataset_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: AuthenticatedUser = Depends(get_current_user),
) -> DatasetOut:
    result = await db.execute(
        select(Dataset).where(Dataset.id == dataset_id, Dataset.is_active == True)  # noqa: E712
    )
    dataset = result.scalar_one_or_none()
    if dataset is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dataset not found")
    return dataset  # type: ignore[return-value]


# ── GET /datasets/{id}/preview ────────────────────────────────────────────────

@router.get("/{dataset_id}/preview", response_model=PreviewResponse)
async def preview_dataset(
    dataset_id: uuid.UUID,
    limit: int = Query(default=20, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: AuthenticatedUser = Depends(get_current_user),
) -> PreviewResponse:
    # Verify dataset exists
    result = await db.execute(
        select(Dataset).where(Dataset.id == dataset_id, Dataset.is_active == True)  # noqa: E712
    )
    dataset = result.scalar_one_or_none()
    if dataset is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dataset not found")

    table_name = f"dataset_{dataset_id.hex}"
    try:
        conn = get_duckdb()
        rel = conn.execute(f'SELECT * FROM "{table_name}" LIMIT ?', [limit])
        columns = [desc[0] for desc in rel.description]
        rows = [list(row) for row in rel.fetchall()]
        conn.close()
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to query dataset: {exc}",
        )

    return PreviewResponse(
        columns=columns,
        rows=rows,
        total_rows=dataset.row_count,
    )


# ── GET /datasets/{id}/field-values ──────────────────────────────────────────

@router.get("/{dataset_id}/field-values", response_model=List[str])
async def get_field_values(
    dataset_id: uuid.UUID,
    field: str = Query(..., description="Column name to get distinct values for"),
    db: AsyncSession = Depends(get_db),
    current_user: AuthenticatedUser = Depends(get_current_user),
) -> List[str]:
    result = await db.execute(
        select(Dataset).where(Dataset.id == dataset_id, Dataset.is_active == True)  # noqa: E712
    )
    dataset = result.scalars().first()
    if dataset is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dataset not found")

    # Validate field exists in schema
    schema_columns = [c["name"] for c in dataset.schema_info.get("columns", [])]
    if field not in schema_columns:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Field '{field}' not found")

    table_name = f"dataset_{dataset_id.hex}"
    try:
        conn = get_duckdb()
        try:
            rel = conn.execute(
                f'SELECT DISTINCT "{field}" FROM "{table_name}" WHERE "{field}" IS NOT NULL ORDER BY "{field}" LIMIT 200'
            )
        except Exception:
            # ORDER BY may fail on mixed-type columns; retry without it
            rel = conn.execute(
                f'SELECT DISTINCT "{field}" FROM "{table_name}" WHERE "{field}" IS NOT NULL LIMIT 200'
            )
        values = [str(row[0]) for row in rel.fetchall()]
        conn.close()
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(exc))

    return values


# ── Auto permission config ────────────────────────────────────────────────────

# 字段名 → (角色, user属性) 映射
_PARTNER_RLS_FIELDS = {
    "伙伴运营负责人": ("ops_manager",     "user.name"),
    "伙伴销售负责人": ("partner_manager",  "user.name"),
}

async def _auto_configure_permissions(
    db: AsyncSession,
    dataset_id: uuid.UUID,
    schema: DatasetSchema,
) -> None:
    """上传后自动检测伙伴相关字段，配置访问授权和 RLS 规则。"""
    from app.models.permission import DatasetAccess, RlsRule

    col_names = {c.name for c in schema.columns}
    matched = {f: v for f, v in _PARTNER_RLS_FIELDS.items() if f in col_names}
    if not matched:
        return  # 无伙伴字段，跳过

    # 全角色授权（admin/analyst 全量，ops_manager/partner_manager 走 RLS）
    for role in ("admin", "analyst", "ops_manager", "partner_manager"):
        db.add(DatasetAccess(
            id=uuid.uuid4(),
            dataset_id=dataset_id,
            grantee_type="role",
            grantee_id=role,
            access_level="read",
        ))

    # 对匹配字段配 RLS
    for field, (role, value_source) in matched.items():
        db.add(RlsRule(
            id=uuid.uuid4(),
            dataset_id=dataset_id,
            name=f"{role}_{field}过滤",
            condition_type="attribute_match",
            field=field,
            operator="eq",
            value_source=value_source,
            applies_to_roles=[role],
            exempt_roles=["admin", "analyst"],
            is_active=True,
        ))

    await db.commit()


# ── GET /datasets/{id}/download ───────────────────────────────────────────────

@router.get("/{dataset_id}/download")
async def download_dataset(
    dataset_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: AuthenticatedUser = Depends(get_current_user),
) -> StreamingResponse:
    """下载数据集为 Excel，管理员上传的数据自动套 RLS 权限过滤。"""
    result = await db.execute(
        select(Dataset).where(Dataset.id == dataset_id, Dataset.is_active == True)  # noqa: E712
    )
    dataset = result.scalar_one_or_none()
    if dataset is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dataset not found")

    # 权限检查：非 admin 只能下载自己上传的，或有访问授权的数据集
    from app.models.permission import DatasetAccess
    is_owner = str(dataset.created_by) == str(current_user.user_id)
    if not is_owner and current_user.role != "admin":
        access = await db.execute(
            select(DatasetAccess).where(
                DatasetAccess.dataset_id == dataset_id,
                DatasetAccess.grantee_type == "role",
                DatasetAccess.grantee_id == current_user.role,
            )
        )
        if access.scalars().first() is None:
            raise HTTPException(status_code=403, detail="无权下载此数据集")

    # 取数据，对有 RLS 的数据集注入权限过滤
    from app.services.permission_resolver import PermissionResolver
    from app.services.query_executor import execute_query

    if dataset.source_type == "duckdb_view" and dataset.file_path:
        table_name = dataset.file_path
    else:
        table_name = f"dataset_{dataset_id.hex}"

    base_sql = f'SELECT * FROM "{table_name}"'

    resolver = PermissionResolver(db)
    try:
        resolved = await resolver.resolve(
            user=current_user,
            dataset_id=dataset_id,
            base_sql=base_sql,
            table_name=table_name,
        )
        final_sql = resolved["sql"]
    except HTTPException:
        raise
    except Exception:
        final_sql = base_sql

    try:
        qr = execute_query(final_sql, dataset_table=table_name)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"数据查询失败: {exc}")

    # 生成 Excel
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = dataset.name[:31]  # Excel sheet 名最长 31 字符

    # 表头
    header_fill = PatternFill(start_color="2D3142", end_color="2D3142", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col_idx, col_name in enumerate(qr.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(12, len(str(col_name)) + 4)

    # 数据行
    for row_idx, row in enumerate(qr.rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    safe_name = dataset.name.replace("/", "_").replace("\\", "_")
    encoded = quote(f"{safe_name}.xlsx", safe="")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


# ── POST /datasets/{id}/restore ───────────────────────────────────────────────

@router.post("/{dataset_id}/restore", status_code=status.HTTP_200_OK)
async def restore_dataset(
    dataset_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: AuthenticatedUser = Depends(get_current_user),
) -> dict:
    """从回收站恢复数据集（仅 admin）。"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可恢复数据集")
    result = await db.execute(
        select(Dataset).where(Dataset.id == dataset_id, Dataset.is_active == False)  # noqa: E712
    )
    dataset = result.scalar_one_or_none()
    if dataset is None:
        raise HTTPException(status_code=404, detail="Dataset not found in trash")
    dataset.is_active = True
    await db.commit()
    return {"id": str(dataset_id), "restored": True}


# ── DELETE /datasets/{id} ─────────────────────────────────────────────────────

@router.delete("/{dataset_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_dataset(
    dataset_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: AuthenticatedUser = Depends(get_current_user),
) -> None:
    """删除数据集（仅 admin）。"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可删除数据集")

    result = await db.execute(
        select(Dataset).where(Dataset.id == dataset_id, Dataset.is_active == True)  # noqa: E712
    )
    dataset = result.scalar_one_or_none()
    if dataset is None:
        raise HTTPException(status_code=404, detail="Dataset not found")

    # Soft-delete only — keep DuckDB table and file so restore works
    dataset.is_active = False
    await db.commit()
