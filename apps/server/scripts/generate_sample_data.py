"""
Generate examples/sample_partner_data.xlsx — 100 rows of fake partner sales data.

Usage (from repo root):
    python apps/server/scripts/generate_sample_data.py
"""
from __future__ import annotations

import random
import sys
from pathlib import Path

# Allow running from any directory
ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT / "apps" / "server"))

import openpyxl
from openpyxl.styles import Font, PatternFill

REGIONS = ["East", "North", "South", "West"]
TIERS = ["Gold", "Silver", "Bronze"]
PRODUCT_LINES = ["Enterprise", "SMB", "Startup", "Education"]
MONTHS = [
    "2025-07", "2025-08", "2025-09", "2025-10", "2025-11", "2025-12",
    "2026-01", "2026-02", "2026-03",
]

PARTNER_POOL = [
    ("TechCorp", "P001"), ("DataSystems", "P002"), ("CloudBase", "P003"),
    ("InfoTech", "P004"), ("NetSolutions", "P005"), ("DigiPartner", "P006"),
    ("AlphaGroup", "P007"), ("BetaWorks", "P008"), ("GammaSoft", "P009"),
    ("DeltaNet", "P010"), ("EpsilonTech", "P011"), ("ZetaCloud", "P012"),
]

random.seed(42)

rows = []
for i in range(100):
    partner_name, partner_id = PARTNER_POOL[i % len(PARTNER_POOL)]
    region = random.choice(REGIONS)
    tier = random.choice(TIERS)
    month = MONTHS[i % len(MONTHS)]
    product_line = random.choice(PRODUCT_LINES)

    # Revenue: Gold > Silver > Bronze
    base_revenue = {"Gold": 80_000, "Silver": 40_000, "Bronze": 15_000}[tier]
    revenue = round(base_revenue * (0.7 + random.random() * 0.6), 2)

    # Deal count: 1–20
    deal_count = random.randint(1, 20)

    rows.append([partner_name, partner_id, region, tier, month, revenue, deal_count, product_line])

# Write workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "PartnerSales"

headers = ["partner_name", "partner_id", "region", "tier", "month", "revenue", "deal_count", "product_line"]

# Header row with style
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font

# Data rows
for row_idx, row in enumerate(rows, start=2):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Column widths
for col, width in zip("ABCDEFGH", [20, 10, 10, 10, 12, 14, 12, 14]):
    ws.column_dimensions[col].width = width

out_path = ROOT / "examples" / "sample_partner_data.xlsx"
out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(out_path))
print(f"Created {out_path}  ({len(rows)} rows)")
